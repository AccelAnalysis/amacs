#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from amacs_io import DATASET_ORDER, ROOT, all_datasets

BLACK = '0B0B0D'
GOLD = 'D6A23A'
IVORY = 'F7F3EA'
GRAPHITE = '252932'
WHITE = 'FFFFFF'
LINE = 'DDD6C8'


def flatten(value: Any) -> Any:
    if value is None:
        return ''
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False, separators=(',', ':'))
    return value


def sheet_title(dataset: str) -> str:
    preferred = {
        'concept_properties': 'Concept Properties',
        'property_values': 'Property Values',
        'request_families': 'Request Families',
        'requirement_types': 'Requirement Types',
        'requirement_bundles': 'Requirement Bundles',
        'response_sections': 'Response Sections',
        'response_templates': 'Response Templates',
        'decision_factors': 'Decision Factors',
        'decision_templates': 'Decision Templates',
        'governance_profiles': 'Governance Profiles',
        'readiness_rules': 'Readiness Rules',
    }
    return preferred.get(dataset, dataset.replace('_', ' ').title())[:31]


def main() -> None:
    parser = argparse.ArgumentParser(description='Generate a non-authoritative AMACS review workbook.')
    parser.add_argument('--output', required=True)
    args = parser.parse_args()
    target = Path(args.output)
    target.parent.mkdir(parents=True, exist_ok=True)

    data = all_datasets(ROOT)
    datasets = [(name, data[name]) for name in DATASET_ORDER if data[name]]

    workbook = Workbook()
    workbook.remove(workbook.active)

    notice = workbook.create_sheet('NOTICE')
    notice.sheet_view.showGridLines = False
    notice['A1'] = 'AMACS generated review copy'
    notice['A1'].font = Font(bold=True, size=20, color=BLACK)
    notice['A1'].fill = PatternFill('solid', fgColor=IVORY)
    notice['A3'] = 'This workbook is generated from the canonical Git source. Changes made here do not modify AMACS.'
    notice['A3'].font = Font(size=12, color=GRAPHITE)
    notice['A3'].alignment = Alignment(wrap_text=True)
    notice['A5'] = f"Release: {(ROOT / 'VERSION').read_text(encoding='utf-8').strip()}"
    notice['A6'] = 'Authority: Git source → pull request → validation → approved release'
    notice['A8'] = 'Dataset'
    notice['B8'] = 'Records'
    for cell in notice[8]:
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill('solid', fgColor=BLACK)
    for row_index, (dataset, records) in enumerate(datasets, start=9):
        notice.cell(row=row_index, column=1, value=dataset)
        notice.cell(row=row_index, column=2, value=len(records))
    notice.column_dimensions['A'].width = 55
    notice.column_dimensions['B'].width = 16
    notice.freeze_panes = 'A9'

    thin = Side(style='thin', color=LINE)
    for dataset, records in datasets:
        worksheet = workbook.create_sheet(sheet_title(dataset))
        worksheet.sheet_view.showGridLines = False
        columns: list[str] = []
        for record in records:
            for key in record:
                if key not in columns:
                    columns.append(key)
        worksheet.append(columns)
        for cell in worksheet[1]:
            cell.font = Font(bold=True, color=WHITE)
            cell.fill = PatternFill('solid', fgColor=GRAPHITE)
            cell.alignment = Alignment(wrap_text=True, vertical='center')
            cell.border = Border(bottom=thin)
        worksheet.row_dimensions[1].height = 34
        for row_number, record in enumerate(records, start=2):
            worksheet.append([flatten(record.get(key)) for key in columns])
            if row_number % 2 == 0:
                for cell in worksheet[row_number]:
                    cell.fill = PatternFill('solid', fgColor='FFFCF5')
        worksheet.freeze_panes = 'A2'
        worksheet.auto_filter.ref = worksheet.dimensions
        for index, key in enumerate(columns, start=1):
            max_len = max(len(str(worksheet.cell(row=row, column=index).value or ''))
                          for row in range(1, min(worksheet.max_row, 300) + 1))
            cap = 55 if key in {'definition', 'description', 'purpose', 'message_template', 'rationale', 'items'} else 36
            worksheet.column_dimensions[get_column_letter(index)].width = min(max(max_len + 2, 12), cap)
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical='top', wrap_text=True)

    workbook.save(target)
    print(target)


if __name__ == '__main__':
    main()
